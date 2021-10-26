import http.client
import hashlib
import urllib
import random
import json
import xlrd
import os
import sys
import time
from xlutils.copy import copy 
import re
import openpyxl


def baidu_trans(q,form,to):
    appid = ''  # 填写你的appid
    secretKey = ''  # 填写你的密钥
    httpClient = None
    myurl = '/api/trans/vip/translate'
    fromLang = form   #原文语种
    toLang = to   #译文语种
    salt = random.randint(32768, 65536)
    sign = appid + q + str(salt) + secretKey
    sign = hashlib.md5(sign.encode()).hexdigest()
    myurl = myurl + '?appid=' + appid + '&q=' + urllib.parse.quote(q) + '&from=' + fromLang + '&to=' + toLang + '&salt=' + str(salt) + '&sign=' + sign

    try:
        httpClient = http.client.HTTPConnection('api.fanyi.baidu.com')
        httpClient.request('GET', myurl)

        # response是HTTPResponse对象
        response = httpClient.getresponse()
        result_all = response.read().decode("utf-8")
        result = json.loads(result_all)
        return result

    except Exception as e:
        print(e)
    finally:
        if httpClient:
            httpClient.close()


def txt_write(path,File_name,form,to):
    result_name = input ("请输入要保存的文件名(包括文件类型)\n:")
    #f读取txt文件，fn写txt文件 
    f = open(path+"\\"+File_name, "r", encoding='utf-8')
    fn = open(path+"\\"+result_name,"a", encoding='utf-8')
    lines = f.readlines()
    #预处理
    count = 0
    word = ''
    length = 0
    #遍历内容并拼接
    for line in lines: 
        #最后一行内容，用作标记，确保把最后不够3000的字符也翻译并写入
        last_line = lines[-1]
        #空行跳过处理
        if line in ['\n','\r\n']:
            pass
        elif line.strip() == "":
            pass
        else:
            #拼接需要翻译的单词
            word = word +line  
            count = count + 1
            #长度作为判断标准
            length = len(word) 
            #写入最后的不够3000的字符
            if line == last_line:
                trans = baidu_trans(word,form,to)
                time.sleep(1)
                for i in range(0,10000):
                    fn.write(trans['trans_result'][i]['dst'])
                    fn.write('\n')
                    word = ''
                    if i == count-1:
                        count = 0
                        break 
            #写入字符
            if length >= 3000:
                trans = baidu_trans(word,form,to)
                time.sleep(1)
                for i in range(0,10000):
                    #返回的是json格式，所以可以通过遍历来达到效果
                    print(trans['trans_result'][i]['dst'])
                    fn.write(trans['trans_result'][i]['dst'])
                    fn.write('\n')
                    word = ''
                    if i == count-1:
                        count = 0
                        break  
    print('OK!')
    fn.close()
    f.close()

def excel_write(path,File_name,form,to):
    File_path = path + '\\' + File_name
    #加载已经存在的excel
    wk = openpyxl.load_workbook(File_path)  
    wk_name = wk.sheetnames
    #选择sheet表处理
    wk_sheet = wk[wk_name[0]]
    maxrow = wk_sheet.max_row
    print('最大行：%s行'%maxrow)

    columns_flag = input("请输入需要翻译的列(1,2,3,4...):")
    columns_flag_int = int(columns_flag)
    columns_result = input("请输入需要储存的列(1,2,3,4...):")
    columns_result_int =int(columns_result)
    #预处理
    num_flag = 1
    sum_content = ' '
    Total = 1
    number = 1
    number_flag = 1
    Uncle_number=[1]*100000 #作为是否翻译和是否需要翻译的标记
    #预处理，先遍历，把
    for row_flag in range(1,maxrow+1):
        unit_content = wk_sheet.cell(row=row_flag,column = columns_flag_int).value
        if unit_content == None:
            Uncle_number[num_flag] = 0
        num_flag+=1
   
    #遍历每一行
    for row_flag in range(1,maxrow+1):
        unit_content = wk_sheet.cell(row=row_flag,column = columns_flag_int).value
        #对空内容进行标记处理
        if unit_content == None:
            Uncle_number[number] = 0
            #因为可能最大值会在判断为空的里面，所以需要判断进而对不够3000字符的进行处理
            if number == (maxrow):
                trans_unit_content = baidu_trans(sum_content,form,to)
                time.sleep(1)
                #遍历拼接好的返回的json的内容
                for j in range(0,Total - number_flag+1):
                    for z in range(1,maxrow):
                        if Uncle_number[z] == 1:
                            #对返回的内容做处理
                            translate = trans_unit_content['trans_result'][j]['dst']
                            translate = translate.replace('"','')
                            translate = translate.replace('“','')
                            translate = translate.replace('”','')
                            #对空数据进行判断
                            if len(translate)==0:
                                print(z)
                                break
                            else:
                                print(z)
                                print(translate)
                                wk_sheet.cell(row=z, column=columns_result_int, value=translate)
                                Uncle_number[z] = 0
                                break 
                #作为标记，来计算拼接数量和进行对应翻译行的标记  
                # 把翻译拼接完的数据清空                     
                    z+=1
                sum_content = ''
            number += 1     
        else:
            #对数据进行预处理，即进行拼接
            sum_content = sum_content + unit_content + '  "\n"  ' 
            #计算长度，方便后面判断
            length = len(sum_content)
            #到最后可能会有不够的字符，也要进行处理
            if number == (maxrow):
                trans_unit_content = baidu_trans(sum_content,form,to)
                time.sleep(1)
                #对这些字符依次处理 Total -number_flag+1 为拼接字符的数量
                for j in range(0,Total - number_flag+1):
                    for z in range(1,maxrow):
                        if Uncle_number[z] == 1:
                            #对翻译的数据进行清洗处理
                            translate = trans_unit_content['trans_result'][j]['dst']
                            translate = translate.replace('"','')
                            translate = translate.replace('“','')
                            translate = translate.replace('”','')
                            
                            if len(translate)==0:
                                print(z)
                                break
                            else:
                                print(z)
                                print(translate)
                                wk_sheet.cell(row=z, column=columns_result_int, value=translate)
                                Uncle_number[z] = 0
                                break
                    z+=1
                sum_content = ''
            if length >=1000:
                #对大于1000的字符进行处理
                trans_unit_content = baidu_trans(sum_content,form,to)
                time.sleep(1)
                for j in range(0,Total - number_flag+1):
                    for z in range(1,maxrow):
                        if Uncle_number[z] == 1:
                      
                            translate = trans_unit_content['trans_result'][j]['dst']
                            translate = translate.replace('"','')
                            translate = translate.replace('“','')
                            translate = translate.replace('”','')
                           
                            if len(translate)==0:
                                print(z)
                                break
                            else:
                                print(z)
                                print(translate)
                                wk_sheet.cell(row=z, column=columns_result_int, value=translate)
                                Uncle_number[z] = 0
                                break  
                
            #作为标记，来计算拼接数量和进行对应翻译行的标记  
            # 把翻译拼接完的数据清空               
                    z+=1
                num_flag = number  
                sum_content = ''
                number_flag = Total
            
            Total += 1
            number += 1
    print('OK!')
    wk.save(File_path)  


    
    


if __name__ == '__main__':

    path = os.path.split(os.path.realpath(__file__))[0]
    print(path)
    #返回的是字符型的
    flag = input("请输入需要翻译的语言\n 1 英译汉\n 2 汉译英\n 3 其它语言译汉\n:")
    #对翻译语言进行处理
    if flag == '1':
        form = 'en'
        to = 'zh'
    elif flag == '2':
        form = 'zh'
        to = 'en'
    elif flag == '3':
        form = 'auto'
        to = 'zh'
    else:
        print('没有这一选项')
    #对文件进行查找处理
    File_Type = input ("请输入需要处理的文件类型\n 1 txt文件\n 2 excel文件\n:")
    File_name = input ("请输入处理的文件名(包括文件类型)\n:")
    if File_Type == '1':
        txt_write(path,File_name,form,to)
        
    elif File_Type == '2':
        excel_write(path,File_name,form,to)   

    else:
        print('没有这一选项')